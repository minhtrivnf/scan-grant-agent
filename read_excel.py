import openpyxl
wb = openpyxl.load_workbook(r"D:\VNF\scan-grant-agent\output\runs\20260811_111825\Grant_Scan_Tracker_RetriV_VNF.xlsx", data_only=True)
ws = wb["🗄️ Database"]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
row2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
for h, v in zip(headers, row2):
    print(f"{h}: {v}")
